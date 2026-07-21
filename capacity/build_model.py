"""Generates fde_capacity_model.xlsx. Row references are tracked by key, never
typed by hand, so inserting a line can't silently break a formula."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

BLUE  = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10)
BOLD  = Font(name='Arial', size=10, bold=True)
HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE = Font(name='Arial', size=13, bold=True, color='0B4DA2')
NOTE  = Font(name='Arial', size=9, italic=True, color='5B6770')
UNIT  = Font(name='Arial', size=9, color='5B6770')
FILLH = PatternFill('solid', fgColor='0B4DA2')
FILLIN= PatternFill('solid', fgColor='FFFFCC')
_t = Side(style='thin', color='D7DEE7')
BOX = Border(left=_t, right=_t, top=_t, bottom=_t)


class Sheet:
    def __init__(self, wb, name, title, sub):
        self.ws = wb.create_sheet(name) if wb.sheetnames != ['Sheet'] or wb['Sheet'].max_row > 1 \
                  else wb.active
        self.ws.title = name
        self.r = 1
        self.ref = {}
        self.ws['A1'] = title; self.ws['A1'].font = TITLE
        self.ws['A2'] = sub;   self.ws['A2'].font = NOTE
        for col, w in (('A',48), ('B',16), ('C',11), ('D',64)):
            self.ws.column_dimensions[col].width = w
        self.ws.sheet_view.showGridLines = False
        self.r = 4

    def section(self, text):
        for col in range(1, 5):
            c = self.ws.cell(self.r, col, text if col == 1 else None)
            c.font = HDR; c.fill = FILLH
        self.r += 1

    def gap(self):
        self.r += 1

    def note(self, text, bold=False):
        self.ws.cell(self.r, 1, text).font = BOLD if bold else NOTE
        self.r += 1

    def _put(self, key, label, value, unit, note, font, fill, fmt):
        self.ws.cell(self.r, 1, label).font = BLACK
        c = self.ws.cell(self.r, 2, value)
        c.font = font; c.border = BOX
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        self.ws.cell(self.r, 3, unit).font = UNIT
        self.ws.cell(self.r, 4, note).font = NOTE
        if key: self.ref[key] = f'B{self.r}'
        self.r += 1

    def inp(self, key, label, value, unit='', note='', fmt=None):
        self._put(key, label, value, unit, note, BLUE, FILLIN, fmt)

    def calc(self, key, label, formula, unit='', note='', fmt=None):
        self._put(key, label, formula, unit, note, BLACK, None, fmt)

    def f(self, template):
        """Substitute {key} with the tracked cell address."""
        return '=' + template.format(**self.ref)


wb = Workbook()

# ===================== 1. GPU SIZING =====================
s = Sheet(wb, 'GPU Sizing', 'Worked Model 1 — GPU sizing for a self-hosted LLM',
          'Yellow cells are inputs. Everything else is a formula. Change an input and the sheet moves.')
s.section('DEMAND')
s.inp('users', 'Peak concurrent users', 500, 'users', 'Peak, not average. Size for the busy hour.', '#,##0')
s.inp('rpu',   'Requests per user per busy hour', 12, 'req/hr', 'From the observed workflow, not from hope.', '#,##0')
s.calc('rps',  'Peak requests per second', s.f('{users}*{rpu}/3600'), 'RPS', 'Demand converted to a rate.', '0.00')
s.inp('outtok','Output tokens per response (mean)', 350, 'tok', 'Measure this. Do not assume it.', '#,##0')
s.inp('intok', 'Input tokens per request (mean)', 2400, 'tok', 'RAG context dominates input length.', '#,##0')
s.calc('tps',  'Peak output tokens per second', s.f('{rps}*{outtok}'), 'tok/s', 'The number that sizes the fleet.', '#,##0')
s.gap()
s.section('HARDWARE THROUGHPUT')
s.inp('gputps','Decode throughput per GPU (batched)', 2200, 'tok/s', 'Benchmark YOUR model on YOUR GPU. Vendor figures are a ceiling.', '#,##0')
s.inp('util',  'Target utilisation ceiling', 0.70, '', 'Above ~70% queueing delay dominates tail latency.', '0%')
s.calc('eff',  'Effective throughput per GPU', s.f('{gputps}*{util}'), 'tok/s', '', '#,##0')
s.calc('gcomp','GPUs needed for decode', s.f('CEILING({tps}/{eff},1)'), 'GPU', '', '0')
s.gap()
s.section('KV CACHE MEMORY CHECK')
s.inp('d',     'Model hidden dim (d)', 8192, '', 'Llama-70B class.', '#,##0')
s.inp('layers','Layers', 80, '', '', '#,##0')
s.inp('qh',    'Attention heads', 64, '', '', '#,##0')
s.inp('kvh',   'KV heads (GQA)', 8, '', 'Set equal to attention heads for old MHA models. Getting this wrong overstates KV memory 8x.', '#,##0')
s.inp('bytes', 'Bytes per element (fp16=2, fp8=1)', 2, 'B', 'Quantising the KV cache is the cheapest capacity win available.', '0')
s.inp('ctx',   'Max context length', 8000, 'tok', '', '#,##0')
s.calc('dkv',  'KV dimension per layer', s.f('{d}*{kvh}/{qh}'), '', 'GQA shrinks this by the head ratio.', '#,##0')
s.calc('kvtok','KV bytes per token', s.f('2*{layers}*{dkv}*{bytes}'), 'B/tok', '2 (K and V) x layers x KV dim x bytes.', '#,##0')
s.calc('kvseq','KV bytes per full-context sequence', s.f('{kvtok}*{ctx}'), 'B', '', '#,##0')
s.calc('kvgb', 'GB per full-context sequence', s.f('{kvseq}/1024^3'), 'GB', '', '0.00')
s.calc('conc', 'Expected in-flight requests', s.f('{rps}*3'), 'req', 'Little\'s Law: arrival rate x ~3s mean service time.', '0.0')
s.inp('gpumem','GPU memory', 80, 'GB', 'H100 80GB.', '#,##0')
s.inp('wmem',  'Memory taken by weights', 140, 'GB', '70B at fp16 — spans two GPUs before any KV cache exists.', '#,##0')
s.calc('gmem', 'GPUs to hold weights (tensor parallel)', s.f('CEILING({wmem}/{gpumem},1)'), 'GPU', '', '0')
s.calc('kvfree','Free memory for KV across those GPUs', s.f('{gmem}*{gpumem}-{wmem}'), 'GB', '', '0.0')
s.calc('seqs', 'Concurrent full-context sequences supported',
       s.f('IF({kvgb}=0,0,FLOOR({kvfree}/{kvgb},1))'), 'seq',
       'Compare against expected in-flight requests above.', '0')
s.calc('kvok', 'KV headroom check',
       s.f('IF({seqs}>={conc},"OK — compute-bound","KV-BOUND: add GPUs, shorten context, or quantise the cache")'), '',
       'When this trips, KV cache is your real limit — not FLOPs. Most teams discover it in production.')
s.gap()
s.section('RESULT')
s.calc('gpus', 'GPUs required (max of compute and memory paths)', s.f('MAX({gcomp},{gmem})'), 'GPU', '', '0')
s.calc('gpuha','Replicas for HA (N+1)', s.f('{gpus}+1'), 'GPU', 'Never size a customer deployment at N.', '0')
s.inp('gpuhr', 'GPU $/hour (on-demand)', 4.50, '$/hr', 'Committed-use pricing cuts this materially.', '$#,##0.00')
s.calc('gpumo','Monthly GPU cost', s.f('{gpuha}*{gpuhr}*730'), '$', '730 hours in an average month.', '$#,##0')
s.gap()
s.note('Assumption: decode-bound. If input tokens greatly exceed output tokens, model prefill separately —')
s.note('prefill is compute-bound and can dominate time-to-first-token even when decode capacity is fine.')

# ===================== 2. RAG COST =====================
s = Sheet(wb, 'RAG Cost per 1K', 'Worked Model 2 — Cost per 1,000 RAG queries',
          'The number customers actually ask for. Itemised so you can defend every line of it.')
s.section('PER-QUERY SHAPE')
s.inp('intok',  'Input tokens (system + context + question)', 2400, 'tok', '', '#,##0')
s.inp('outtok', 'Output tokens', 350, 'tok', '', '#,##0')
s.inp('qemb',   'Query embedding tokens', 40, 'tok', '', '#,##0')
s.inp('rerank', 'Rerank candidates per query', 25, 'doc', 'Top-25 into the cross-encoder, top-5 into context.', '#,##0')
s.gap()
s.section('UNIT PRICES — re-validate against vendor pricing pages before quoting')
s.inp('pin',   'Generation input $/M tokens', 3.00, '$/M', 'Mid-tier band. See the cost-band cheat sheet.', '$#,##0.00')
s.inp('pout',  'Generation output $/M tokens', 15.00, '$/M', '', '$#,##0.00')
s.inp('pemb',  'Embedding $/M tokens', 0.13, '$/M', '', '$#,##0.00')
s.inp('prr',   'Rerank $/1K calls', 2.00, '$/1K', 'Priced per search call (up to ~100 docs), not per document.', '$#,##0.00')
s.inp('vdb',   'Vector DB $/month', 700, '$', 'Amortised across monthly volume below.', '$#,##0')
s.inp('vol',   'Queries per month', 400000, 'q', '', '#,##0')
s.gap()
s.section('COST PER 1,000 QUERIES — NO CACHING')
s.calc('cin',  'Generation input', s.f('1000*{intok}/1000000*{pin}'), '$', '', '$#,##0.0000')
s.calc('cout', 'Generation output', s.f('1000*{outtok}/1000000*{pout}'), '$', '', '$#,##0.0000')
s.calc('cemb', 'Query embedding', s.f('1000*{qemb}/1000000*{pemb}'), '$', '', '$#,##0.0000')
s.calc('crr',  'Reranking', s.f('{prr}'), '$', 'One call per query, so $/1K calls is already the per-1K cost.', '$#,##0.0000')
s.calc('cvdb', 'Vector DB (amortised)', s.f('IF({vol}=0,0,{vdb}/{vol}*1000)'), '$', '', '$#,##0.0000')
s.calc('sub',  'Subtotal', s.f('{cin}+{cout}+{cemb}+{crr}+{cvdb}'), '$/1K', '', '$#,##0.0000')
s.gap()
s.section('CACHE SENSITIVITY')
s.inp('hit',   'Prompt-cache hit rate', 0.45, '', 'Stable system prompt plus stable retrieved chunks.', '0%')
s.inp('disc',  'Cached-input discount', 0.90, '', 'Typical provider discount on a cached prefix.', '0%')
s.inp('sem',   'Semantic output-cache hit rate', 0.15, '', 'High in support, near zero in research work.', '0%')
s.calc('eff',  'Effective cost per 1,000 queries',
       s.f('({cin}*(1-{hit}*{disc})+{cout}+{cemb}+{crr}+{cvdb})*(1-{sem})'), '$/1K', '', '$#,##0.0000')
s.calc('per1', 'Cost per single query', s.f('{eff}/1000'), '$', '', '$#,##0.000000')
s.calc('mo',   'Monthly run rate at stated volume', s.f('{eff}*{vol}/1000'), '$', '', '$#,##0')
s.calc('save', 'Saving from caching', s.f('({sub}-{eff})/{sub}'), '', 'Caching is the highest-return line in this model.', '0.0%')

# ===================== 3. VOICE LATENCY =====================
s = Sheet(wb, 'Voice Latency', 'Worked Model 3 — The sub-800ms voice budget',
          'A latency target you cannot decompose is a wish. This is the decomposition.')
s.section('BUDGET LINES (p95, milliseconds)')
s.inp('nin',  'Network in (client to edge)', 40, 'ms', '')
s.inp('vad',  'Endpointing / silence detection', 180, 'ms', 'The biggest lever, and the one teams forget.')
s.inp('asr',  'Streaming ASR finalisation', 90, 'ms', 'Partials arrive earlier; this is the final commit.')
s.inp('ret',  'Retrieval (if grounded)', 110, 'ms', 'Cache aggressively or take it off the hot path.')
s.inp('ttft', 'LLM time to first token', 220, 'ms', 'Small model by default; escalate only on low confidence.')
s.inp('tts',  'TTS time to first audio', 90, 'ms', 'Streaming TTS. Non-streaming blows the budget on its own.')
s.inp('nout', 'Network out (edge to client)', 40, 'ms', '')
s.inp('jit',  'Jitter buffer', 30, 'ms', '')
s.calc('tot', 'Total perceived latency', s.f('{nin}+{vad}+{asr}+{ret}+{ttft}+{tts}+{nout}+{jit}'), 'ms', '', '#,##0')
s.inp('target','Target', 800, 'ms', 'Past ~800ms callers begin talking over the agent.', '#,##0')
s.calc('head', 'Headroom', s.f('{target}-{tot}'), 'ms', 'Negative means redesign, not tuning.', '#,##0')
s.calc('verd', 'Verdict', s.f('IF({head}>=0,"Within budget","OVER BUDGET — redesign")'), '')
s.gap()
s.note('Levers, in order of payoff:', bold=True)
for t in ('1. Tune the VAD. Often 80–120ms back for no quality loss.',
          '2. Start TTS on the first sentence, not the finished response.',
          '3. Pre-fetch retrieval on the partial transcript, off the hot path.',
          '4. Route to a small model by default; escalate mid-turn only when confidence is low.',
          '5. Co-locate ASR, LLM and TTS in one region. Cross-region alone can cost 100ms+.'):
    s.note(t)

# ===================== 4. SELF-HOST BREAK-EVEN =====================
s = Sheet(wb, 'Self-host Breakeven', 'Worked Model 4 — Self-hosting vs API break-even',
          'The answer is rarely "self-hosting is cheaper". It is "cheaper above X tokens/month", and X is large.')
s.section('API PATH')
s.inp('price', 'Blended API $/M tokens', 6.00, '$/M', 'Weighted across input and output.', '$#,##0.00')
s.inp('vol',   'Monthly tokens (millions)', 400, 'M', '', '#,##0')
s.calc('api',  'Monthly API cost', s.f('{price}*{vol}'), '$', '', '$#,##0')
s.gap()
s.section('SELF-HOST PATH')
s.inp('gpus',  'GPUs required', 4, 'GPU', 'Carry this across from the GPU Sizing sheet.', '0')
s.inp('gpuhr', 'GPU $/hour (committed)', 2.80, '$/hr', 'Committed-use pricing, not on-demand.', '$#,##0.00')
s.calc('ghw',  'Monthly GPU cost', s.f('{gpus}*{gpuhr}*730'), '$', '', '$#,##0')
s.inp('fte',   'Engineer FTE on the serving stack', 0.4, 'FTE', 'The line everyone omits. vLLM does not operate itself.', '0.0')
s.inp('sal',   'Fully-loaded engineer $/year', 260000, '$', '', '$#,##0')
s.calc('gppl', 'Monthly people cost', s.f('{fte}*{sal}/12'), '$', '', '$#,##0')
s.calc('self', 'Monthly self-host total', s.f('{ghw}+{gppl}'), '$', '', '$#,##0')
s.gap()
s.section('DECISION')
s.calc('save', 'Monthly saving from self-hosting', s.f('{api}-{self}'), '$', 'Negative means the API is cheaper.', '$#,##0')
s.calc('be',   'Break-even volume (M tokens/month)', s.f('IF({price}=0,0,{self}/{price})'), 'M',
       'Below this, self-hosting loses on pure cost.', '#,##0')
s.calc('verd', 'Verdict on cost alone', s.f('IF({save}>0,"Self-host wins","API wins")'), '')
s.gap()
s.note('Cost is one axis and usually not the deciding one. Data residency, weight control, and')
s.note('air-gap requirements decide most real self-hosting conversations before price is discussed.')

# ===================== 5. BUSINESS VALUE =====================
s = Sheet(wb, 'Business Value', 'Worked Model 5 — Tokens to dollars to hours saved',
          'The conversion that wins the executive readout. Run it before the pilot, not after.')
s.section('BASELINE — measure it, do not assume it')
s.inp('users','Users in scope', 900, 'people', '', '#,##0')
s.inp('tasks','Task instances per user per day', 14, '', '', '#,##0')
s.inp('before','Minutes per task, before', 9.0, 'min', 'From a timed audit, not from a survey.', '0.0')
s.inp('rate', 'Fully-loaded cost per user-hour', 68, '$', '', '$#,##0')
s.inp('days', 'Working days per month', 21, 'd', '', '0')
s.gap()
s.section('AFTER')
s.inp('after', 'Minutes per task, after', 6.2, 'min', 'Same timed audit, post-pilot.', '0.0')
s.inp('adopt', 'Adoption rate', 0.55, '',
      'The number that actually determines ROI. Most models quietly set it to 100%.', '0%')
s.gap()
s.section('RESULT')
s.calc('saved','Minutes saved per task', s.f('{before}-{after}'), 'min', '', '0.0')
s.calc('hours','Hours saved per month', s.f('{users}*{tasks}*{days}*{saved}*{adopt}/60'), 'hr', '', '#,##0')
s.calc('gross','Gross monthly value (time-equivalent)', s.f('{hours}*{rate}'), '$', 'What a naive model stops at.', '$#,##0')
s.inp('real', 'Realisation rate', 0.30, '',
      'Fraction of saved time that becomes real capacity or cost reduction. 20-40% is honest; 100% is vendor math.', '0%')
s.calc('rvalue','Realised monthly value', s.f('{gross}*{real}'), '$', 'The number a CFO will accept.', '$#,##0')
s.inp('run',   'Monthly run cost', 3800, '$', 'Carry across from the RAG Cost sheet.', '$#,##0')
s.calc('net',  'Net monthly value', s.f('{rvalue}-{run}'), '$', '', '$#,##0')
s.inp('build', 'One-off build cost', 400000, '$', '', '$#,##0')
s.calc('pay',  'Payback period', s.f('IF({net}<=0,"never",{build}/{net})'), 'months', '', '0.0')
s.gap()
s.note('Two multipliers decide this model: adoption and realisation. Halve either and payback doubles.')
s.note('Neither is a modelling detail — they are the whole argument. Model quality barely moves this sheet,')
s.note('which is why adoption is the metric an FDE defends in the quarterly review.')

wb.save('fde_capacity_model.xlsx')
print('written')
